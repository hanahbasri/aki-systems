import shutil
from pathlib import Path
from openpyxl import load_workbook

from aki_engine import AKIInput, ProductLine, CapexInput, calculate_aki, N_YEARS

TEMPLATE_PATH = Path(__file__).parent / "template_aki.xlsx"
YEAR_COLS = ["E", "F", "G", "H", "I"]  # Tahun 1-5


def generate_excel(result: dict, output_path: str) -> str:
    shutil.copy(str(TEMPLATE_PATH), output_path)
    wb = load_workbook(output_path)
    inp: AKIInput = result["input"]
    months = result["months_per_year"]  # [m1, m2, m3, m4, m5]

    _fill_sales_revenue(wb, inp, months)
    _fill_investasi(wb, inp, months)
    _fill_cogs_conn(wb, inp)
    _fill_opex_pct(wb, inp)
    _fill_cover_manual(wb, inp)

    wb.save(output_path)
    return output_path


# ── Sales & Revenue (master input sheet) ─────────────────────────────────────
def _fill_sales_revenue(wb, inp: AKIInput, months: list):
    ws = wb["Sales & Revenue"]

    # B2: Nama Program
    ws["B2"] = inp.nama_program

    # E4: Start year (tahun pertama billing)
    ws["E4"] = inp.rencana_selesai

    # E6:I6: Jumlah bulan aktif per tahun
    for i, col in enumerate(YEAR_COLS):
        ws[f"{col}6"] = months[i]

    # ── Bersihkan baris produk lama (B9:K30) ──────────────────────────────
    for row in range(9, 31):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            cell = ws[f"{col}{row}"]
            # Jangan hapus formula
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    # ── Pisah produk Butuh JT vs Tanpa JT ────────────────────────────────
    jt_prods = [p for p in inp.products if p.tipe == "Butuh JT"]
    non_jt_prods = [p for p in inp.products if p.tipe != "Butuh JT"]

    # Section header
    ws["B8"] = "Layanan Butuh JT"
    ws["B11"] = "Layanan tidak perlu JT"

    # ── Isi produk Butuh JT (baris 9, 10) ────────────────────────────────
    # Template hanya punya 1 row Butuh JT (row 9).
    # Kita isi row 9 untuk produk pertama JT, row 10 dst jika ada lebih.
    # Subscriber qty diasumsikan sama sepanjang kontrak (copy dari col E).
    for i, prod in enumerate(jt_prods[:2]):  # max 2 baris Butuh JT di template
        row = 9 + i
        ws[f"B{row}"] = prod.name
        ws[f"C{row}"] = prod.satuan
        ws[f"D{row}"] = inp.lokasi
        ws[f"E{row}"] = prod.qty
        # F:I = formula =E9 (copy qty dari tahun 1)
        for col in ["F", "G", "H", "I"]:
            ws[f"{col}{row}"] = f"=E{row}"

    # ── Isi produk Tanpa JT (baris 12..16) ───────────────────────────────
    for i, prod in enumerate(non_jt_prods[:5]):
        row = 12 + i
        ws[f"B{row}"] = prod.name
        ws[f"C{row}"] = prod.satuan
        ws[f"D{row}"] = inp.lokasi
        ws[f"E{row}"] = prod.qty
        for col in ["F", "G", "H", "I"]:
            ws[f"{col}{row}"] = f"=E{row}"

    # ── Revenue section (baris 18 dst) ───────────────────────────────────
    ws["B18"] = "Revenue "
    ws["B19"] = "=B8"   # "Layanan Butuh JT"

    # Baris 20..21 = revenue Butuh JT (formula: C*qty*bulan)
    for i, prod in enumerate(jt_prods[:2]):
        rev_row = 20 + i
        sub_row = 9 + i   # baris subscriber qty
        ws[f"B{rev_row}"] = f"=B{sub_row}"
        ws[f"C{rev_row}"] = prod.monthly_price
        for j, col in enumerate(YEAR_COLS):
            ws[f"{col}{rev_row}"] = f"=C{rev_row}*{col}{sub_row}*${col}$6"

        # OTC: simpan di kolom M (tarif OTC)
        ws[f"J{rev_row}"] = prod.otc_price  # OTC per unit (kolom cadangan)

    # "Layanan tidak perlu JT" header di baris 22
    ws["B22"] = "=B11"

    # Baris 23..27 = revenue non-JT
    for i, prod in enumerate(non_jt_prods[:5]):
        rev_row = 23 + i
        sub_row = 12 + i
        ws[f"B{rev_row}"] = f"=B{sub_row}"
        ws[f"C{rev_row}"] = prod.monthly_price
        for j, col in enumerate(YEAR_COLS):
            ws[f"{col}{rev_row}"] = f"=C{rev_row}*{col}{sub_row}*${col}$6"

    # Revenue summary row 32
    ws["B32"] = "Revenue"
    for col in YEAR_COLS:
        ws[f"{col}32"] = f"=SUM({col}20:{col}30)"

    # Total Revenue row 33
    ws["B33"] = "Total Revenue"
    ws["K33"] = "=SUM(E32:J32)"

    # OTC total di baris terpisah (PSB/instalasi) — row 21 jika produk JT ada OTC
    # Ini dibutuhkan oleh 4.Dir sheet yang reference row 15 di Sales&Revenue
    # untuk OTC instalasi
    # Kita tulis OTC instalasi di row 21 (baris "PSB" default template)
    total_otc_price = 0
    for prod in jt_prods:
        total_otc_price += prod.otc_price * prod.qty
    if total_otc_price > 0:
        ws["C41"] = total_otc_price  # tarif OTC (digunakan oleh formula di 4.Dir row 14)


# ── 8. Investasi & depresiasi ─────────────────────────────────────────────────
def _fill_investasi(wb, inp: AKIInput, months: list):
    ws = wb["8. Investasi & depresiasi"]

    if not inp.capex:
        return

    # B2: Nama program (sudah ada formula =Sales&Revenue!B2)
    # G20 = Total CAPEX dari regional table — ini yang direferensikan Cpx!L49

    # Regional table (row 17):
    ws["B17"] = f"REG"
    ws["C17"] = inp.id_lop or inp.lokasi
    ws["D17"] = 0          # ID IHLD (kosong)
    ws["E17"] = inp.capex.material
    ws["F17"] = inp.capex.jasa
    ws["G17"] = "=E17+F17"

    # G20 = Total CAPEX (referenced by 3.Cpx L49)
    ws["G20"] = "=SUM(G17:G19)"

    # Row 7: Investasi New
    ws["C7"] = "Investasi New"
    ws["D7"] = "JT FO"
    ws["E7"] = "Paket"
    ws["F7"] = 1
    # G7 = =G20 (sudah ada formula di template, tapi kita pastikan)
    ws["G7"] = "=G20"
    ws["H7"] = inp.capex.lifetime_years

    # I3, K3, L3, M3 = months per year (untuk depresiasi proporsional)
    # I3=month yr1, K3=yr3, L3=yr4, M3=yr5 (template skip J3=yr2, pakai F6)
    # Isi manual karena formula template reference ke '[131]Sales & Revenue' (file lama)
    ws["I3"] = months[0]
    ws["K3"] = months[2]
    ws["L3"] = months[3]
    ws["M3"] = months[4]

    # Row 8: BOP Lakwas (formula =G7*0.4% sudah ada di template)
    ws["C8"] = "BOP Lakwas"
    ws["D8"] = "Support"
    ws["E8"] = "Lot"
    ws["F8"] = 1
    ws["H8"] = inp.capex.lifetime_years


# ── COGS Conn ─────────────────────────────────────────────────────────────────
def _fill_cogs_conn(wb, inp: AKIInput):
    """
    Isi tarif MRC dan OTC di COGS Conn.
    Formula di sheet ini: G7 = (I7*(1-0.3)) = 70% dari tarif MRC
    I7 = Tarif Dasar MRC (bulanan)
    J7 = Tarif Dasar OTC (instalasi)
    """
    ws = wb["COGS Conn"]

    # Bersihkan baris 7..25
    for row in range(7, 26):
        for col in ["I", "J", "K", "L", "M", "N", "O"]:
            cell = ws[f"{col}{row}"]
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    # Produk Butuh JT (rows 7+)
    jt_prods = [p for p in inp.products if p.tipe == "Butuh JT"]
    non_jt_prods = [p for p in inp.products if p.tipe != "Butuh JT"]

    row = 7
    for prod in jt_prods[:2]:
        ws[f"I{row}"] = prod.monthly_price   # Tarif Dasar MRC
        ws[f"J{row}"] = prod.otc_price       # Tarif Dasar OTC
        # Formula COGS sudah ada: F7=J7*(1-0.25), G7=(I7*(1-0.3))
        row += 1

    # Non-JT mulai row 9 (setelah section header row 8)
    row = 9
    for prod in non_jt_prods[:10]:
        ws[f"I{row}"] = prod.monthly_price
        ws[f"J{row}"] = prod.otc_price
        row += 1


# ── 5. Opx — isi % O&M di assumptions ────────────────────────────────────────
def _fill_opex_pct(wb, inp: AKIInput):
    ws = wb["5. Opx"]
    # D65: % O&M per bulan (sudah ada di template default 0.12)
    ws["D65"] = inp.om_pct
    ws["D64"] = 0   # BHP = 0 (per catatan template)


# ── 1. Cover — isi field yang tidak punya formula ─────────────────────────────
def _fill_cover_manual(wb, inp: AKIInput):
    ws = wb["1. Cover"]

    # Tahun di judul
    ws["B2"] = f"HASIL EVALUASI AKI CAPEX  PT. TELKOM  TAHUN  {inp.rencana_selesai}"

    # Field yang tidak ada formula-nya di template
    ws["E7"] = inp.jumlah_lop          # Jumlah LOP
    ws["E8"] = inp.id_lop or ""        # ID LOP
    ws["E9"] = inp.lokasi              # Lokasi
    ws["E10"] = len(inp.products)      # Volume (jumlah produk/SSL)
    ws["E11"] = "Lokasi"              # Satuan
    ws["E17"] = 1                      # Waktu Penyelesaian (bulan)
    ws["E20"] = inp.teknologi          # Teknologi


# ── Test ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    test_input = AKIInput(
        nama_program="Penyediaan Layanan Jaringan PT QL Agrofood Tanjungsari",
        nama_customer="PT QL Agrofood",
        lokasi="CAU",
        cust_group="Large Enterprise",
        teknologi="FO",
        rencana_selesai=2026,
        jumlah_lop=1,
        id_lop="11270075-PT3-CAU-FC-CIBADAK",
        kontrak_tahun=2,
        kontrak_bulan=0,
        start_month=1,
        products=[
            ProductLine("Astinet Reguler 150 Mbps @ 1 SSL", qty=1,
                        monthly_price=9106000, otc_price=2500000,
                        is_hsi=False, tipe="Butuh JT"),
            ProductLine("Last Mile Astinet", qty=1,
                        monthly_price=750000, otc_price=0,
                        is_hsi=False, tipe="Butuh JT"),
            ProductLine("Astinet Reguler 300 Mbps @ 1 SSL", qty=1,
                        monthly_price=16964667, otc_price=0,
                        is_hsi=False, tipe="Tanpa JT"),
            ProductLine("Astinet Lite 10 Mbps @ 1 SSL", qty=1,
                        monthly_price=535000, otc_price=0,
                        is_hsi=False, tipe="Tanpa JT"),
            ProductLine("Astinet Lite 100 Mbps @ 1 SSL", qty=1,
                        monthly_price=6141000, otc_price=0,
                        is_hsi=False, tipe="Tanpa JT"),
        ],
        capex=CapexInput(material=40637662, jasa=43469314, lifetime_years=5),
        om_pct=0.12,
    )

    result = calculate_aki(test_input)
    out = "/home/claude/test_v2_output.xlsx"
    generate_excel(result, out)
    print(f"✅ Generated: {out}")

    # Verify key cells
    from openpyxl import load_workbook as lw
    wb2 = lw(out)
    sr = wb2["Sales & Revenue"]
    print(f"B2  = {sr['B2'].value}")
    print(f"E4  = {sr['E4'].value}")
    print(f"E6:I6 months = {[sr[f'{c}6'].value for c in 'EFGHI']}")
    print(f"B9  = {sr['B9'].value}")
    print(f"C20 = {sr['C20'].value}  (monthly price prod 1)")
    print(f"E20 = {sr['E20'].value}  (formula revenue yr1)")

    inv = wb2["8. Investasi & depresiasi"]
    print(f"E17 material = {inv['E17'].value}")
    print(f"F17 jasa     = {inv['F17'].value}")
    print(f"G17 formula  = {inv['G17'].value}")
    print(f"G7  formula  = {inv['G7'].value}")